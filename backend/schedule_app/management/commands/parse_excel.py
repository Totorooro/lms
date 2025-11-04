from django.core.management.base import BaseCommand
from datetime import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
from schedule_app.models import Subject, Group, Lesson

class Command(BaseCommand):
    help = 'Parse Excel schedule and populate the database'

    def handle(self, *args, **options):
        wb = load_workbook('C:/Users/555/Desktop/test/1.xlsx') # path
        ws = wb["1к-ИСАУ (1) "] # Excel Sheet -_-


        # days in nums for DB
        days = {
            'понедельник': 1,
            'вторник': 2,
            'среда': 3,
            'четверг': 4,
            'пятница': 5,
            'суббота': 6,
        }

        # types for DB
        types_map = {
            'лекция': 'lecture',
            'лабораторная работа': 'lab',
            'семинар': 'seminar',
            'ДОТ': 'practice',  
        }

        # num of lesson to time
        pair_times = {
            1: ('09:00:00', '10:30:00'),
            2: ('10:40:00', '12:10:00'),
            3: ('12:50:00', '14:20:00'),
            4: ('14:30:00', '16:00:00'),
            5: ('16:10:00', '17:40:00'),
            6: ('17:50:00', '19:20:00'),
        }

        # choose group to parse
        group_name = '1013'

        # get or create group
        group, _ = Group.objects.get_or_create(name=group_name)

        # Checks whether the cell is in the combined
        def get_merged_cell_value(ws, row, col):
            for merged_range in ws.merged_cells.ranges:
                min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
                if min_row <= row <= max_row and min_col <= col <= max_col:
                    return ws.cell(row=min_row, column=min_col).value # take the leftmost value of the cell
            return None

        current_pair = None
        current_day = None
        group_col = 5  # Column for group(3 is 1011 etc.)

        def parse_lesson(text, ws, main_row, col):
            text = text.strip() 

            if not text:
                return {'room': '', 'teacher': '', 'subject': '', 'type': ''}

            words = re.split(r'\s+', text) # split text to parse


            room_parts = []
            i = 0

            if i < len(words) and re.match(r'^\d+-\d+$', words[i]): # chech for room_parts(1-120)
                room_parts.append(words[i])
                i += 1
            elif i < len(words) and words[i] == 'с/к':
                room_parts.append(words[i])
                i += 1
                if i < len(words) and words[i] == 'Олимп':
                    room_parts.append(words[i])
                    i += 1
            elif i < len(words) and words[i] == 'корпус':
                room_parts.append(words[i])
                i += 1
                if i < len(words):
                    room_parts.append(words[i])  
                    i += 1
                if i < len(words) and words[i] == '(конгресс-центр)':
                    room_parts.append(words[i])
                    i += 1

            room = ' '.join(room_parts)
            words = words[i:] # deleting processed words
            titles = ['проф.', 'доц.', 'доцент', 'ст.преп.', 'с.п.', 'пр.', 'профессор', 'проф', 'доц']
            teacher_parts = []
            skip_len = 0
            if words and re.match(r'[А-Я]\.[А-Я]\.$', words[-1]): # checking for initials
                teacher_parts = [words[-1]] 
                skip_len += 1
                if len(words) > 1:
                    teacher_parts.insert(0, words[-2])
                    skip_len += 1
                    if len(words) > 2 and any(title in words[-3].lower() for title in titles):
                        skip_len += 1  

            teacher = ' '.join(teacher_parts)
            subject_words = words[:-skip_len] if skip_len > 0 else words
            subject = ' '.join(subject_words)


            subject = re.sub(r'\s*\([^)]*\)$', '', subject).strip()


            cell = ws.cell(row=main_row, column=col)
            lesson_type = 'семинар' # default value
            if room.startswith('ДОТ'): # determining the type of item
                lesson_type = 'ДОТ'
            elif cell.font.italic:
                lesson_type = 'лабораторная работа'
            elif any(title in ' '.join(words).lower() for title in titles):
                lesson_type = 'лекция'

            return {'room': room, 'teacher': teacher, 'subject': subject, 'type': lesson_type}

        def save_to_db(parsed, current_day, current_pair, week):
            subject_name = parsed['subject'].strip()
            if not subject_name:
                return  


            subject, _ = Subject.objects.get_or_create(name=subject_name)

            week_type = 'every' if not week else ('odd' if week == 'нечетная' else 'even')

            parsed_type = parsed['type']
            lesson_type = types_map.get(parsed_type, 'seminar') 

            day_num = days.get(current_day.lower())
            if day_num is None:
                return 


            pair_int = int(current_pair)
            if pair_int not in pair_times:
                return  
            start_time_str, end_time_str = pair_times[pair_int]


            start_time_obj = time.fromisoformat(start_time_str)
            end_time_obj = time.fromisoformat(end_time_str)

   
            Lesson.objects.create(
                subject=subject,
                type=lesson_type,
                start_time=start_time_obj,
                end_time=end_time_obj,
                day=day_num,
                classroom=parsed['room'].strip(),
                teacher=parsed['teacher'].strip() if parsed['teacher'] else None,
                group=group,
                week_type=week_type,
            )

        for row in range(5, 68):  
            day_cell = ws['A' + str(row)].value
            if day_cell and isinstance(day_cell, str) and day_cell.strip().lower() in ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота']:
                current_day = day_cell.strip()

            if current_day is None:
                continue

            pair_cell = ws['B' + str(row)].value

            if pair_cell and str(pair_cell).isdigit(): 
                current_pair = pair_cell

                merged_value = get_merged_cell_value(ws, row, group_col) # Gets the value from the merged cell or default value
                if merged_value is not None:
                    main_text = merged_value
                else:
                    main_text = ws[get_column_letter(group_col) + str(row)].value or ''

                additional_text = ''
                original_row = row  
                next_row = row + 1
                if next_row <= ws.max_row:
                    next_pair = ws['B' + str(next_row)].value
                    if next_pair is None or not str(next_pair).isdigit():
                        merged_value_next = get_merged_cell_value(ws, next_row, group_col)
                        if merged_value_next is not None:
                            additional_text = merged_value_next
                        else:
                            additional_text = ws[get_column_letter(group_col) + str(next_row)].value or ''
                        row = next_row  
                if additional_text and main_text and additional_text in main_text:
                    full_text = main_text
                else:
                    full_text = (main_text + ' ' + additional_text).strip()
                if not full_text:
                    continue 
                 

                full_text = re.sub(r'(\d+)\s*-\s*(\d+)', r'\1-\2', full_text)
                full_text = re.sub(r'\s+', ' ', full_text).strip()

                full_text = full_text.replace('с/к', 'с-к')

                full_text = re.sub(r'\s*/+\s*', ' / ', full_text)
                full_text = full_text.replace('с-к', 'с/к')
                is_odd_only = False
                is_even_only = False
                if full_text.startswith(' / '):
                    is_even_only = True
                    full_text = full_text[3:].strip()
                if full_text.endswith(' / '):
                    is_odd_only = True
                    full_text = full_text[:-3].strip()
                parts = full_text.split(' / ')
                if len(parts) > 1:
                    for idx, part in enumerate(parts):
                        if part.strip():
                            week = 'нечетная' if idx == 0 else 'четная'
                            type_row = original_row if idx == 0 else (next_row if additional_text else original_row)
                            parsed = parse_lesson(part.strip(), ws, type_row, group_col)
                            save_to_db(parsed, current_day, current_pair, week)
                else:
                    week = 'нечетная' if is_odd_only else ('четная' if is_even_only else '')
                    parsed = parse_lesson(full_text, ws, original_row, group_col)
                    save_to_db(parsed, current_day, current_pair, week)

        self.stdout.write(self.style.SUCCESS('Data has been added to the database.'))